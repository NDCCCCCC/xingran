#!/usr/bin/env python3
"""
工位管理 - 部门映射表增加"父级部门链路"列

方案说明:
1. 读取用户提供的部门映射(dept_name, dept_code)
2. 通过 PostgreSQL sys_dept.ancestors 字段递归解析父级链路
3. 写出新的 Excel,新增"父级部门链路"列(用 " → " 分隔,用于在 Excel 中区分同名部门)

运行前置:
  pip install psycopg2-binary openpyxl pandas
"""

import argparse
import sys
import os
from pathlib import Path

try:
    import psycopg2
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError as e:
    print(f"缺少依赖: {e}", file=sys.stderr)
    print("请运行: pip install psycopg2-binary openpyxl pandas", file=sys.stderr)
    sys.exit(1)


# ── 数据源 dept_list (用户提供) ──────────────────────────────────────────────
# 也可换成读取已有 xlsx 第 1 列/第 2 列,见 --from-xlsx
DEFAULT_DEPT_LIST = [
    ("车商业务部",                    "004"),
    ("电网销业务部",                  "00A"),
    ("个人营销业务销售部",            "00B"),
    ("个人营销业务销售部",            "00C"),
    ("个人营销业务销售部",            "00D"),
    ("个人营销业务销售部",            "00E"),
    ("个人营销业务销售部",            "00F"),
    ("个人营销业务销售部",            "00G"),
    ("个人营销业务销售一部",          "00H"),
    ("个客中心线上平台部个人营销业务销售部", "00I"),
    ("个人营销业务销售部",            "00J"),
    ("个人营销业务销售部",            "00K"),
    ("个人营销业务销售部",            "00L"),
    ("个人营销业务销售部",            "00M"),
    ("个人营销业务销售部",            "00N"),
    ("赤壁支公司",                    "00O"),
    ("竹山支公司",                    "00P"),
]

# 模拟数据(在没有 Postgres 时使用,展示输出格式)
# 说明:每个 dept_code 对应的父级链路,按 " → " 分隔
MOCK_PARENT_MAP = {
    "004": "中国人民财产保险股份有限公司 → 湖北省分公司 → 武汉市分公司 → 车商业务部",
    "00A": "中国人民财产保险股份有限公司 → 湖北省分公司 → 武汉市分公司 → 电网销业务部",
    "00B": "中国人民财产保险股份有限公司 → 湖北省分公司 → 咸宁中心支公司 → 个人营销业务销售部",
    "00C": "中国人民财产保险股份有限公司 → 湖北省分公司 → 宜昌中心支公司 → 个人营销业务销售部",
    "00D": "中国人民财产保险股份有限公司 → 湖北省分公司 → 襄阳中心支公司 → 个人营销业务销售部",
    "00E": "中国人民财产保险股份有限公司 → 湖北省分公司 → 黄石中心支公司 → 个人营销业务销售部",
    "00F": "中国人民财产保险股份有限公司 → 湖北省分公司 → 十堰中心支公司 → 个人营销业务销售部",
    "00G": "中国人民财产保险股份有限公司 → 湖北省分公司 → 荆州中心支公司 → 个人营销业务销售部",
    "00H": "中国人民财产保险股份有限公司 → 湖北省分公司 → 鄂州中心支公司 → 个人营销业务销售一部",
    "00I": "中国人民财产保险股份有限公司 → 湖北省分公司 → 个客中心线上平台部 → 个人营销业务销售部",
    "00J": "中国人民财产保险股份有限公司 → 湖北省分公司 → 荆门中心支公司 → 个人营销业务销售部",
    "00K": "中国人民财产保险股份有限公司 → 湖北省分公司 → 孝感中心支公司 → 个人营销业务销售部",
    "00L": "中国人民财产保险股份有限公司 → 湖北省分公司 → 黄冈中心支公司 → 个人营销业务销售部",
    "00M": "中国人民财产保险股份有限公司 → 湖北省分公司 → 咸宁中心支公司 → 通城支公司 → 个人营销业务销售部",
    "00N": "中国人民财产保险股份有限公司 → 湖北省分公司 → 随州中心支公司 → 个人营销业务销售部",
    "00O": "中国人民财产保险股份有限公司 → 湖北省分公司 → 咸宁中心支公司 → 赤壁支公司",
    "00P": "中国人民财产保险股份有限公司 → 湖北省分公司 → 十堰中心支公司 → 竹山支公司",
}


# ── SQL: 从 PostgreSQL sys_dept 解析父级链路 ───────────────────────────────
PARENT_CHAIN_SQL = """
WITH RECURSIVE dept_ancestors AS (
    -- 起点:目标部门本身
    SELECT
        d.id,
        d.dept_code,
        d.dept_name,
        d.parent_id,
        0 AS depth,
        d.dept_name::TEXT AS path
    FROM sys_dept d
    WHERE d.dept_code = ANY(%(codes)s)

    UNION ALL

    -- 递归:向上找父级
    SELECT
        p.id,
        p.dept_code,
        p.dept_name,
        p.parent_id,
        da.depth + 1,
        p.dept_name || ' → ' || da.path
    FROM sys_dept p
    INNER JOIN dept_ancestors da ON p.id = da.parent_id
)
SELECT
    da.dept_code          AS dept_code,
    da.dept_name          AS dept_name,
    da.path               AS parent_chain
FROM dept_ancestors da
WHERE da.depth = (
    -- 只保留每个 dept_code 的最深(即自身)那条记录
    SELECT MAX(depth) FROM dept_ancestors WHERE dept_code = da.dept_code
)
ORDER BY da.dept_code;
"""


def fetch_parent_chains_from_db(conn_str: str, dept_codes: list[str]) -> dict[str, str]:
    """连接 PostgreSQL,返回 {dept_code: parent_chain}"""
    conn = psycopg2.connect(conn_str)
    try:
        with conn.cursor() as cur:
            cur.execute(PARENT_CHAIN_SQL, {"codes": dept_codes})
            rows = cur.fetchall()
            return {code: chain for code, _, chain in rows}
    finally:
        conn.close()


def build_workbook(dept_list: list[tuple[str, str]],
                   parent_map: dict[str, str],
                   out_path: Path,
                   source_label: str) -> None:
    """写出新的 xlsx,带父级链路列(同名部门因此可区分)"""
    wb = Workbook()
    ws = wb.active
    ws.title = "部门映射(去重)"

    # 表头
    headers = ["部门编码 (dept_code)", "部门名称 (dept_name)", "父级部门链路 (parent_chain)"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF", name="Arial")
        c.fill = PatternFill("solid", start_color="305496")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = Border(bottom=Side(style="medium"))

    # 数据
    thin = Side(style="thin", color="BFBFBF")
    for r, (name, code) in enumerate(dept_list, start=2):
        chain = parent_map.get(code, "(未匹配,请检查 dept_code 是否已导入 sys_dept 表)")
        for col, val in enumerate([code, name, chain], start=1):
            c = ws.cell(row=r, column=col, value=val)
            c.font = Font(name="Arial")
            c.alignment = Alignment(vertical="center", wrap_text=(col == 3))
            c.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # 列宽
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 90

    # 冻结表头
    ws.freeze_panes = "A2"

    # 元数据 sheet
    meta = wb.create_sheet("说明")
    meta["A1"] = "数据源"
    meta["A1"].font = Font(bold=True, name="Arial")
    meta["B1"] = source_label
    meta["A2"] = "生成时间"
    meta["A2"].font = Font(bold=True, name="Arial")
    meta["B2"] = str(__import__("datetime").datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    meta["A3"] = "解析说明"
    meta["A3"].font = Font(bold=True, name="Arial")
    meta["B3"] = "用 sys_dept.parent_id 自连接递归得到完整链路, 用 ' → ' 分隔."

    wb.save(out_path)
    print(f"已生成: {out_path}")


def load_from_xlsx(xlsx_path: Path) -> list[tuple[str, str]]:
    """从已有 Excel 读 dept_name/dept_code 两列"""
    wb = load_workbook(xlsx_path)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[0] and row[1]:
            rows.append((str(row[1]).strip(), str(row[0]).strip()))  # (name, code)
    return rows


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--from-xlsx", type=Path, help="从已有 xlsx 读部门列表")
    p.add_argument("--db", default=None,
                   help="PostgreSQL 连接串,例: postgresql://xingran:pwd@localhost:5432/xingran_next")
    p.add_argument("--mock", action="store_true",
                   help="用内置 MOCK_PARENT_MAP 演示输出(无需数据库连接)")
    p.add_argument("--out", type=Path, default=Path("dept_mapping_with_parent.xlsx"),
                   help="输出文件路径")
    args = p.parse_args()

    # 1. 加载部门列表
    if args.from_xlsx:
        dept_list = load_from_xlsx(args.from_xlsx)
        source = f"来自文件 {args.from_xlsx}"
    else:
        dept_list = DEFAULT_DEPT_LIST
        source = "内置默认列表 (用户原始提供的 17 条)"

    # 2. 解析父级链路
    codes = [c for _, c in dept_list]
    if args.mock:
        parent_map = {c: MOCK_PARENT_MAP.get(c, "(MOCK 未覆盖)") for c in codes}
        src_label = source + " | parent from MOCK data"
    elif args.db:
        parent_map = fetch_parent_chains_from_db(args.db, codes)
        src_label = source + f" | DB: {args.db}"
    else:
        print("错误: 必须指定 --db 或 --mock 之一", file=sys.stderr)
        sys.exit(2)

    # 3. 写出
    args.out.parent.mkdir(parents=True, exist_ok=True)
    build_workbook(dept_list, parent_map, args.out, src_label)


if __name__ == "__main__":
    main()
